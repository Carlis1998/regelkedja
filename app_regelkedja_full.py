# app_regelkedja_full.py
# -*- coding: utf-8 -*-
"""
REGELKEDJA – End-to-end Streamlit app (v3)
- Ingest: manuell text, PDF/DOCX/TXT, URL (persistens i SQLite)
- Chunkning: **mycket mer inkluderande** med tre profiler + sammanslagning av för korta block
  * "vagledning" (Får/Get-bilagan) – ett chunk per kontroll-id + underrubrik-split endast när blocket är långt
  * "checklista" (Nöt-checklistan) – ett chunk per kontroll-id
  * "fallback" – stora paragrafklossar (default ~3000 tecken) med överhörning, därefter merge av korta chunks
- Metadata: titel, länk, buckets/tags, nivå (EUF/EUD/SFS/SVF/FÖR/ARÅD/VGL/VERIFIKAT) och art(er)
- Index: KB‑SBERT (KBLab/sentence-bert-swedish-cased) + FAISS (cosine)
- Hybrid-sök: Dense (FAISS) + BM25 (rank_bm25) + *art-belöningslogik* som nu stöder **flera arter**
- Match & Export: läs Excel, bygg fråga per rad, välj bästa per nivå, lägg kandidatkolumner (doc-score + top-3 chunkar)
- Export: behåll originalblad & format, lägg till nya kolumner efter sista kolumn
- Redigering: uppdatera metadata, in-/avaktivera dokument, **re-chunka** dokument med nya inställningar

Kör:
  1) python -m venv .venv && source .venv/bin/activate  # (Windows: .venv\Scripts\activate)
  2) pip install --upgrade pip
  3) pip install -r requirements.txt
  4) streamlit run app_regelkedja_full.py

Kräver: faiss-cpu, PyMuPDF (pymupdf), pdfminer.six, python-docx, trafilatura, rank-bm25, openpyxl, sentence-transformers, pandas, numpy, scikit-learn.
"""

import os, io, re, json, time, math, hashlib, datetime, sqlite3, contextlib
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict, Any, Iterable, Set

import numpy as np
import pandas as pd
import streamlit as st

# Embeddings & retrieval
from sentence_transformers import SentenceTransformer
try:
    import faiss  # type: ignore
except Exception:
    faiss = None

# Sparse retrieval
try:
    from rank_bm25 import BM25Okapi
    HAS_BM25 = True
except Exception:
    HAS_BM25 = False

# File extraction
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except Exception:
    HAS_PYMUPDF = False

try:
    from pdfminer.high_level import extract_text as pdfminer_extract_text
    HAS_PDFMINER = True
except Exception:
    HAS_PDFMINER = False

try:
    import docx  # python-docx
    HAS_DOCX = True
except Exception:
    HAS_DOCX = False

try:
    import trafilatura
    HAS_TRAFILATURA = True
except Exception:
    HAS_TRAFILATURA = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# =========================
# Paths & config
# =========================
APP_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(APP_DIR, "data")
DOC_DIR  = os.path.join(DATA_DIR, "docs")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(DOC_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "registry.sqlite3")
FAISS_PATH = os.path.join(DATA_DIR, "faiss.index")
MODEL_NAME = "KBLab/sentence-bert-swedish-cased"  # KB‑SBERT

# =========================
# Utils
# =========================

def now_iso() -> str:
    return datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def norm_ws(s: str) -> str:
    s = s or ""
    s = s.replace("\r", "\n").replace("\u00AD", "")  # mjukt bindestreck
    s = re.sub(r"-\n", "", s)  # hyfenering över radbryt
    s = re.sub(r"\n\s+\n+", "\n\n", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def sent_split(s: str) -> List[str]:
    return [x.strip() for x in re.split(r"(?<=[\.!?])\s+|\n+|(?<=:)\s+", s) if x.strip()]


def tokens(s: str) -> List[str]:
    s = (s or "").lower()
    s = re.sub(r"[\-\u00AD]", "", s)
    return [t for t in re.split(r"[^a-z0-9åäö]+", s) if t]

# =========================
# DB Schema (+migrering)
# =========================
SCHEMA_SQL = r"""
PRAGMA journal_mode=WAL;
CREATE TABLE IF NOT EXISTS documents (
  doc_id        INTEGER PRIMARY KEY AUTOINCREMENT,
  title         TEXT NOT NULL,
  url           TEXT,
  org           TEXT,
  doc_type      TEXT,
  juridisk_niva TEXT,
  art           TEXT,
  sha256        TEXT,
  file_path     TEXT,
  created_at    TEXT NOT NULL,
  updated_at    TEXT NOT NULL,
  active        INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS buckets (
  bucket_id INTEGER PRIMARY KEY AUTOINCREMENT,
  name      TEXT UNIQUE NOT NULL
);

CREATE TABLE IF NOT EXISTS tags (
  tag_id INTEGER PRIMARY KEY AUTOINCREMENT,
  name   TEXT UNIQUE NOT NULL
);

CREATE TABLE IF NOT EXISTS doc_buckets (
  doc_id    INTEGER NOT NULL,
  bucket_id INTEGER NOT NULL,
  UNIQUE(doc_id, bucket_id)
);

CREATE TABLE IF NOT EXISTS doc_tags (
  doc_id INTEGER NOT NULL,
  tag_id INTEGER NOT NULL,
  UNIQUE(doc_id, tag_id)
);

CREATE TABLE IF NOT EXISTS chunks (
  chunk_id     INTEGER PRIMARY KEY AUTOINCREMENT,
  doc_id       INTEGER NOT NULL,
  ordinal      INTEGER NOT NULL,
  text         TEXT NOT NULL,
  kontrol_id   TEXT,
  section_ref  TEXT,
  page_start   INTEGER,
  page_end     INTEGER,
  anchor_url   TEXT,
  text_hash    TEXT NOT NULL,
  created_at   TEXT NOT NULL,
  updated_at   TEXT NOT NULL,
  active       INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS faiss_map (
  position INTEGER PRIMARY KEY,
  chunk_id INTEGER NOT NULL
);
"""

MIGRATIONS = [
    ("documents", "raw_text", "TEXT"),
]


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_schema():
    with contextlib.closing(get_conn()) as conn:
        conn.executescript(SCHEMA_SQL)
        # add missing columns
        for table, col, typ in MIGRATIONS:
            try:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {typ}")
            except Exception:
                pass
        conn.commit()

# =========================
# Embeddings & FAISS
# =========================
@st.cache_resource(show_spinner=False)
def get_model(name: str = MODEL_NAME) -> SentenceTransformer:
    return SentenceTransformer(name)


class FaissStore:
    def __init__(self, dim: int):
        if faiss is None:
            raise RuntimeError("faiss-cpu saknas (pip install faiss-cpu)")
        self.dim = dim
        self.index = faiss.IndexFlatIP(dim)

    def add(self, vecs: np.ndarray) -> None:
        self.index.add(vecs)

    def search(self, qvecs: np.ndarray, topk: int) -> Tuple[np.ndarray, np.ndarray]:
        return self.index.search(qvecs, topk)

    def ntotal(self) -> int:
        return self.index.ntotal

    def save(self, path: str) -> None:
        faiss.write_index(self.index, path)

    @staticmethod
    def load(path: str) -> "FaissStore":
        idx = faiss.read_index(path)
        obj = FaissStore(idx.d)
        obj.index = idx
        return obj


def normalize_vecs(vecs: np.ndarray) -> np.ndarray:
    n = np.linalg.norm(vecs, axis=1, keepdims=True)
    n = np.maximum(n, 1e-12)
    return (vecs / n).astype(np.float32)


def load_or_new_faiss(model: SentenceTransformer) -> FaissStore:
    dim = model.get_sentence_embedding_dimension()
    if os.path.isfile(FAISS_PATH):
        try:
            store = FaissStore.load(FAISS_PATH)
            if store.dim != dim:
                raise RuntimeError("FAISS-index matchar inte modell-dim; kör Bygg om")
            return store
        except Exception:
            pass
    return FaissStore(dim)


def rebuild_faiss_from_db(model: SentenceTransformer) -> int:
    with contextlib.closing(get_conn()) as conn:
        rows = conn.execute("SELECT chunk_id, text FROM chunks WHERE active=1 ORDER BY chunk_id").fetchall()
    store = load_or_new_faiss(model)
    store.index = faiss.IndexFlatIP(model.get_sentence_embedding_dimension())
    if not rows:
        store.save(FAISS_PATH)
        with contextlib.closing(get_conn()) as conn:
            conn.execute("DELETE FROM faiss_map")
            conn.commit()
        return 0
    texts = [r["text"] for r in rows]
    ids = [int(r["chunk_id"]) for r in rows]
    vecs = get_model().encode(texts, convert_to_numpy=True, show_progress_bar=True)
    vecs = normalize_vecs(vecs)
    store.add(vecs)
    store.save(FAISS_PATH)
    with contextlib.closing(get_conn()) as conn:
        conn.execute("DELETE FROM faiss_map")
        conn.executemany("INSERT INTO faiss_map(position, chunk_id) VALUES (?,?)", list(enumerate(ids)))
        conn.commit()
    return len(ids)


def append_chunks_to_faiss(model: SentenceTransformer, chunk_rows: List[sqlite3.Row]) -> int:
    if not chunk_rows:
        return 0
    store = load_or_new_faiss(model)
    texts = [r["text"] for r in chunk_rows]
    vecs = model.encode(texts, convert_to_numpy=True, show_progress_bar=False)
    vecs = normalize_vecs(vecs)
    start = store.ntotal()
    store.add(vecs)
    store.save(FAISS_PATH)
    with contextlib.closing(get_conn()) as conn:
        conn.executemany("INSERT OR REPLACE INTO faiss_map(position, chunk_id) VALUES (?,?)",
                         [(start+i, int(r["chunk_id"])) for i,r in enumerate(chunk_rows)])
        conn.commit()
    return len(chunk_rows)

# =========================
# Extraction helpers
# =========================

PDF_TITLE_FROM_META = True


def extract_pdf(file_bytes: bytes) -> Tuple[str, List[Tuple[int, str]], Optional[str]]:
    pages: List[Tuple[int, str]] = []
    txt_all, title = "", None
    if HAS_PYMUPDF:
        with fitz.open(stream=file_bytes, filetype="pdf") as doc:
            if PDF_TITLE_FROM_META:
                meta = doc.metadata or {}
                t = (meta.get("title") or meta.get("Title") or "").strip()
                title = t or None
            for i, page in enumerate(doc):
                t = page.get_text("text")
                pages.append((i+1, norm_ws(t)))
                txt_all += t + "\n\n"
    elif HAS_PDFMINER:
        full = pdfminer_extract_text(io.BytesIO(file_bytes))
        txt_all = full or ""
        parts = re.split(r"\f", full or "")
        for i, p in enumerate(parts):
            pages.append((i+1, norm_ws(p)))
    else:
        raise RuntimeError("Installera PyMuPDF eller pdfminer.six för PDF")
    return norm_ws(txt_all), pages, title


def extract_docx(file_bytes: bytes) -> Tuple[str, Optional[str]]:
    if not HAS_DOCX:
        raise RuntimeError("Installera python-docx för DOCX")
    bio = io.BytesIO(file_bytes)
    d = docx.Document(bio)
    title = None
    try:
        core = d.core_properties
        if core and core.title:
            title = core.title.strip() or None
    except Exception:
        pass
    return norm_ws("\n".join(p.text for p in d.paragraphs)), title


def extract_txt(file_bytes: bytes) -> str:
    try:
        return norm_ws(file_bytes.decode("utf-8"))
    except Exception:
        return norm_ws(file_bytes.decode("latin-1", errors="ignore"))


def fetch_url(url: str) -> str:
    if HAS_TRAFILATURA:
        raw = trafilatura.fetch_url(url)
        if not raw:
            return ""
        text = trafilatura.extract(raw, include_links=False, include_formatting=False)
        return norm_ws(text or "")
    try:
        import requests
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        html = r.text
        html = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.I)
        html = re.sub(r"<style[\s\S]*?</style>", " ", html, flags=re.I)
        text = re.sub(r"<[^>]+>", " ", html)
        return norm_ws(text)
    except Exception:
        return ""

# =========================
# Chunking (inkl. efterbearbetning/merge)
# =========================
RE_FARGET = re.compile(r"^(?:FårGet|Får|Get)\s+\d+\.?", re.I)
RE_NOT    = re.compile(r"^Nöt\s+\d+", re.I)
RE_SUB    = re.compile(r"^(?:Kontrollen bör omfatta|Vägledning|Särskild vägledning)", re.I)

RE_SECTION = re.compile(
    r"((?:Art(?:ikel)?\.?\s*\d+[a-z]?)(?:\s*\(\d+\))?|\d+\s*kap\.\s*\d+\s*§[a-z]?|SJVFS\s*\d{4}:\d+|(?:EU|EG)\s*nr\s*\d+\/\d+)",
    re.I
)

@dataclass
class ChunkOptions:
    fallback_max_chars: int = 3000
    fallback_overlap: int = 220
    sub_split_min_chars: int = 2600  # dela först vid underrubriker när blocket är större än detta
    min_chunk_chars: int = 1000      # under denna längd sammanslås med granne
    join_short_chunks: bool = True


def _merge_small_chunks(chunks: List[Dict[str,Any]], min_chars: int) -> List[Dict[str,Any]]:
    if not chunks:
        return []
    out = []
    buf = chunks[0].copy()
    for ch in chunks[1:]:
        if len(buf.get("text","")) < min_chars or (buf.get("kontrol_id") == ch.get("kontrol_id") and len(ch.get("text","")) < min_chars):
            buf["text"] = (buf.get("text","") + "\n\n" + ch.get("text"," ")).strip()
            buf["page_end"] = ch.get("page_end", buf.get("page_end"))
            buf["section_ref"] = buf.get("section_ref") or ch.get("section_ref")
            buf["anchor_url"] = buf.get("anchor_url") or ch.get("anchor_url")
        else:
            out.append(buf)
            buf = ch.copy()
    out.append(buf)
    return out


def chunk_profile_vagledning(text_pages: List[Tuple[int,str]], url: Optional[str], opt: ChunkOptions) -> List[Dict[str,Any]]:
    rows = []
    for pn, page in text_pages:
        for line in page.split("\n"):
            line = line.strip()
            if line:
                rows.append((pn, line))

    chunks: List[Dict[str,Any]] = []
    cur: List[Tuple[int,str]] = []
    cur_id = None
    start_page = None

    def flush():
        nonlocal cur, cur_id, start_page
        if not cur:
            return
        txt = norm_ws("\n".join(x[1] for x in cur))
        sec = RE_SECTION.search(txt)
        ch = {
            "kontrol_id": cur_id,
            "text": txt,
            "section_ref": sec.group(1) if sec else None,
            "page_start": start_page,
            "page_end": cur[-1][0],
            "anchor_url": f"{url}#page={start_page}" if url and start_page else (url or None),
        }
        chunks.append(ch)
        cur, cur_id, start_page = [], None, None

    for pn, line in rows:
        if RE_FARGET.match(line):
            flush()
            cur_id = RE_FARGET.findall(line)[0]
            start_page = pn
            cur = [(pn, line)]
        elif RE_SUB.match(line) and cur_id and len("\n".join(x[1] for x in cur)) > opt.sub_split_min_chars:
            flush()
            cur_id = cur_id
            start_page = pn
            cur = [(pn, line)]
        else:
            cur.append((pn, line))
    flush()
    if opt.join_short_chunks:
        chunks = _merge_small_chunks(chunks, opt.min_chunk_chars)
    return chunks


def chunk_profile_checklista(text_pages: List[Tuple[int,str]], url: Optional[str], opt: ChunkOptions) -> List[Dict[str,Any]]:
    rows = []
    for pn, page in text_pages:
        for line in page.split("\n"):
            line = line.strip()
            if line:
                rows.append((pn, line))
    chunks: List[Dict[str,Any]] = []
    cur: List[Tuple[int,str]] = []
    cur_id = None
    start_page = None

    def flush():
        nonlocal cur, cur_id, start_page
        if not cur:
            return
        txt = norm_ws("\n".join(x[1] for x in cur))
        sec = RE_SECTION.search(txt)
        ch = {
            "kontrol_id": cur_id,
            "text": txt,
            "section_ref": sec.group(1) if sec else None,
            "page_start": start_page,
            "page_end": cur[-1][0],
            "anchor_url": f"{url}#page={start_page}" if url and start_page else (url or None),
        }
        chunks.append(ch)
        cur, cur_id, start_page = [], None, None

    for pn, line in rows:
        if RE_NOT.match(line):
            flush()
            cur_id = RE_NOT.findall(line)[0]
            start_page = pn
            cur = [(pn, line)]
        else:
            cur.append((pn, line))
    flush()
    if opt.join_short_chunks:
        chunks = _merge_small_chunks(chunks, opt.min_chunk_chars)
    return chunks


def chunk_profile_fallback(text: str, url: Optional[str], pages: Optional[List[Tuple[int,str]]], opt: ChunkOptions) -> List[Dict[str,Any]]:
    max_chars, overlap = int(opt.fallback_max_chars), int(opt.fallback_overlap)
    s = norm_ws(text)
    paras = [p.strip() for p in re.split(r"\n\s*\n+", s) if p.strip()]
    chunks: List[Dict[str,Any]] = []
    buf = []
    cur_len = 0

    def make_chunk(content: str) -> Dict[str,Any]:
        pstart = None
        if pages:
            head = content[:120]
            for pn, pg in pages:
                if head and head in pg:
                    pstart = pn
                    break
        sec = RE_SECTION.search(content)
        return {
            "kontrol_id": None,
            "text": content,
            "section_ref": sec.group(1) if sec else None,
            "page_start": pstart,
            "page_end": pstart,
            "anchor_url": f"{url}#page={pstart}" if url and pstart else (url or None),
        }

    for p in paras:
        if cur_len + len(p) + 1 <= max_chars:
            buf.append(p)
            cur_len += len(p) + 1
        else:
            content = norm_ws("\n".join(buf))
            if content:
                chunks.append(make_chunk(content))
            if overlap > 0 and len(content) > overlap:
                tail = content[-overlap:]
                buf = [tail, p]
                cur_len = len(tail) + 1 + len(p)
            else:
                buf = [p]
                cur_len = len(p)
    if buf:
        chunks.append(make_chunk(norm_ws("\n".join(buf))))

    if opt.join_short_chunks:
        chunks = _merge_small_chunks(chunks, opt.min_chunk_chars)
    return chunks

# =========================
# Insert/Rechunk document + chunks
# =========================
LEVEL_CHOICES = ["EUF","EUD","SFS","SVF","FÖR","ARÅD","VGL","VERIFIKAT","ANNAT"]
ARTER_CHOICES = ["Nötkreatur","Får","Getter","Gris"]


def upsert_bucket(conn: sqlite3.Connection, name: str) -> int:
    name = name.strip()
    if not name:
        return 0
    conn.execute("INSERT OR IGNORE INTO buckets(name) VALUES (?)", (name,))
    return int(conn.execute("SELECT bucket_id FROM buckets WHERE name=?", (name,)).fetchone()[0])


def upsert_tag(conn: sqlite3.Connection, name: str) -> int:
    name = name.strip()
    if not name:
        return 0
    conn.execute("INSERT OR IGNORE INTO tags(name) VALUES (?)", (name,))
    return int(conn.execute("SELECT tag_id FROM tags WHERE name=?", (name,)).fetchone()[0])


def _do_chunking(text_all: str, pages: Optional[List[Tuple[int,str]]], url: Optional[str], profile: str, opt: ChunkOptions) -> List[Dict[str,Any]]:
    text_pages = pages or []
    if profile == "auto":
        body = "\n".join([p for _, p in text_pages]) if text_pages else text_all
        count_not = len(re.findall(r"^Nöt\s+\d+", body, flags=re.M))
        count_fg  = len(re.findall(r"^(?:FårGet|Får|Get)\s+\d+", body, flags=re.M))
        profile = "checklista" if count_not >= 3 else ("vagledning" if count_fg >= 3 else "fallback")
    if profile == "vagledning":
        return chunk_profile_vagledning(text_pages, url, opt)
    if profile == "checklista":
        return chunk_profile_checklista(text_pages, url, opt)
    return chunk_profile_fallback(text_all, url, pages, opt)


def insert_document(*, title: str, url: Optional[str], org: Optional[str], doc_type: Optional[str], juridisk_niva: Optional[str], art: Optional[str], file_bytes: Optional[bytes], text_input: Optional[str], buckets: List[str], tags: List[str], chunk_profile: str = "auto", chunk_options: Optional[ChunkOptions] = None) -> int:
    title = (title or "(utan titel)").strip()
    text_all = ""
    pages: Optional[List[Tuple[int,str]]] = None
    src_url = url or None
    file_path = None
    chunk_options = chunk_options or ChunkOptions()

    if file_bytes is not None:
        sha = sha256_bytes(file_bytes)
        ext = ".pdf"
        file_path = os.path.join(DOC_DIR, f"{sha[:16]}{ext}")
        with open(file_path, "wb") as f:
            f.write(file_bytes)
        text_all, pages, pdf_title = extract_pdf(file_bytes)
        if not title and pdf_title:
            title = pdf_title
    elif src_url and not text_input:
        sha = sha256_bytes(src_url.encode("utf-8"))
        text_all = fetch_url(src_url)
    else:
        raw = (text_input or "").encode("utf-8")
        sha = sha256_bytes(raw)
        text_all = text_input or ""

    text_all = norm_ws(text_all)
    if not text_all:
        raise ValueError("Inget textinnehåll hittades")

    chunks = _do_chunking(text_all, pages, src_url, chunk_profile, chunk_options)

    created = updated = now_iso()
    with contextlib.closing(get_conn()) as conn:
        cur = conn.execute(
            """
            INSERT INTO documents(title, url, org, doc_type, juridisk_niva, art, sha256, file_path, created_at, updated_at, active, raw_text)
            VALUES (?,?,?,?,?,?,?,?,?,?,1,?)
            """,
            (title, src_url, org, doc_type, juridisk_niva, art, sha, file_path, created, updated, text_all)
        )
        doc_id = int(cur.lastrowid)
        for b in buckets:
            if b.strip():
                bid = upsert_bucket(conn, b)
                if bid:
                    conn.execute("INSERT OR IGNORE INTO doc_buckets(doc_id, bucket_id) VALUES (?,?)", (doc_id, bid))
        for t in tags:
            if t.strip():
                tid = upsert_tag(conn, t)
                if tid:
                    conn.execute("INSERT OR IGNORE INTO doc_tags(doc_id, tag_id) VALUES (?,?)", (doc_id, tid))

        for i, ch in enumerate(chunks):
            ch_txt = ch["text"]
            ch_hash = sha256_bytes(ch_txt.encode("utf-8"))
            conn.execute(
                """
                INSERT INTO chunks(doc_id, ordinal, text, kontrol_id, section_ref, page_start, page_end, anchor_url, text_hash, created_at, updated_at, active)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,1)
                """,
                (doc_id, i, ch_txt, ch.get("kontrol_id"), ch.get("section_ref"), ch.get("page_start"), ch.get("page_end"), ch.get("anchor_url"), ch_hash, created, updated)
            )
        conn.commit()
        to_embed = conn.execute("SELECT chunk_id, text FROM chunks WHERE doc_id=? AND active=1 ORDER BY chunk_id", (doc_id,)).fetchall()

    append_chunks_to_faiss(get_model(), to_embed)
    return doc_id


def rechunk_document(doc_id: int, *, profile: str, opt: ChunkOptions) -> int:
    with contextlib.closing(get_conn()) as conn:
        row = conn.execute("SELECT url, file_path, raw_text FROM documents WHERE doc_id=?", (doc_id,)).fetchone()
        if not row:
            raise ValueError("Okänt dokument")
        file_path = row["file_path"]
        url = row["url"]
        raw_text = row["raw_text"] or ""
    text_all, pages = "", None
    if file_path and os.path.isfile(file_path):
        with open(file_path, "rb") as f:
            fb = f.read()
        text_all, pages, _ = extract_pdf(fb)
    elif raw_text:
        text_all = raw_text
    else:
        text_all = fetch_url(url or "")
    text_all = norm_ws(text_all)
    if not text_all:
        raise ValueError("Hittade inget textinnehåll för re-chunk")

    chunks = _do_chunking(text_all, pages, url, profile, opt)

    with contextlib.closing(get_conn()) as conn:
        conn.execute("DELETE FROM chunks WHERE doc_id=?", (doc_id,))
        created = updated = now_iso()
        for i, ch in enumerate(chunks):
            conn.execute(
                """
                INSERT INTO chunks(doc_id, ordinal, text, kontrol_id, section_ref, page_start, page_end, anchor_url, text_hash, created_at, updated_at, active)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,1)
                """,
                (
                    doc_id, i, ch["text"], ch.get("kontrol_id"), ch.get("section_ref"),
                    ch.get("page_start"), ch.get("page_end"), ch.get("anchor_url"),
                    sha256_bytes(ch["text"].encode("utf-8")), created, updated
                )
            )
        conn.commit()
    rebuild_faiss_from_db(get_model())
    return len(chunks)

# =========================
# Art-normalisering & matchning (stöd för flera arter)
# =========================
CANON_MAP: Dict[str, Set[str]] = {
    "nötkreatur": {"nötkreatur", "nöt", "nötboskap", "boskap"},
    "får": {"får", "lamm", "får/get", "fårget"},
    "getter": {"get", "getter", "getterna", "getdjur", "får/get", "fårget"},
    "gris": {"gris", "svin"},
}

CANON_KEYS = set(CANON_MAP.keys())


def normalize_art_token(s: str) -> Optional[str]:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    if not s:
        return None
    for canon, alts in CANON_MAP.items():
        if s in alts:
            return canon
    # lite tolerans
    if s.startswith("nöt"):
        return "nötkreatur"
    if s.startswith("get"):
        return "getter"
    if s.startswith("får"):
        return "får"
    if s.startswith("gris") or s.startswith("svin"):
        return "gris"
    return None


def parse_art_list(val: Optional[str]) -> Set[str]:
    if not val:
        return set()
    parts = re.split(r"[,;/\|]", str(val))
    out: Set[str] = set()
    for p in parts:
        canon = normalize_art_token(p)
        if canon:
            out.add(canon)
    return out


def doc_art_set(art_field: Optional[str]) -> Set[str]:
    return parse_art_list(art_field)


def kontrol_id_art_set(kontrol_id: Optional[str]) -> Set[str]:
    s = (kontrol_id or "").lower()
    cand = set()
    if re.search(r"\bnöt", s):
        cand.add("nötkreatur")
    if re.search(r"\bfår", s):
        cand.add("får")
    if re.search(r"\bget", s) or re.search(r"\bgetter", s) or re.search(r"\bfårget", s):
        cand.add("getter")
    if re.search(r"\bgris|\bsvin", s):
        cand.add("gris")
    return cand


def art_sets_match(required: Set[str], doc_set: Set[str], kid_set: Set[str]) -> bool:
    if not required:
        return True
    # godkänn om någon av de efterfrågade arterna finns i dokumentet *eller* kontrol_id
    return len(required & (doc_set | kid_set)) > 0

# =========================
# Search & scoring
# =========================
@dataclass
class Weights:
    w_cos: float = 0.55
    w_bm25: float = 0.25
    w_cov: float = 0.10
    w_len: float = 0.10  # subtract
    bonus_ref: float = 0.12
    bonus_art: float = 0.05


def coverage_score(qtoks: List[str], ttoks: List[str]) -> float:
    if not qtoks:
        return 0.0
    u = set(qtoks)
    s = set(ttoks)
    return len(u & s) / float(len(u))


def length_penalty(q_len: int, t_len: int, hi: float = 1.0, lo: float = 0.35, min_ratio: float = 0.12) -> float:
    if q_len <= 0:
        return 1.0
    ratio = float(t_len) / float(q_len)
    if ratio >= 1.0:
        return hi
    if ratio <= min_ratio:
        return lo
    x = (ratio - min_ratio) / (1.0 - min_ratio)
    return lo + x * (hi - lo)


def detect_refs(s: str) -> List[str]:
    return [m.group(1) for m in RE_SECTION.finditer(s or "")]


def hybrid_search(query: str, *, arter_required: Optional[Set[str]], topk_dense: int = 400, topk_return: int = 120, weights: Weights = Weights()) -> List[Dict[str,Any]]:
    model = get_model()
    vec = model.encode([query], convert_to_numpy=True)
    vec = normalize_vecs(vec)
    store = load_or_new_faiss(model)
    if store.ntotal() == 0:
        return []
    D, I = store.search(vec, min(topk_dense, store.ntotal()))
    pos = I[0].tolist()
    sim = D[0].tolist()

    q_toks = tokens(query)
    q_len = len(query)
    ref_list = detect_refs(query)

    with contextlib.closing(get_conn()) as conn:
        rows = []
        for p, s in zip(pos, sim):
            r = conn.execute(
                """
                SELECT fm.position, fm.chunk_id, c.doc_id, c.text, c.kontrol_id, c.section_ref, c.anchor_url,
                       d.title, d.juridisk_niva, d.doc_type, d.art, d.url
                FROM faiss_map fm
                JOIN chunks c ON c.chunk_id=fm.chunk_id AND c.active=1
                JOIN documents d ON d.doc_id=c.doc_id AND d.active=1
                WHERE fm.position=?
                """,
                (int(p),)
            ).fetchone()
            if not r:
                continue
            # art-gating: kräv träff om arter_required finns
            req = arter_required or set()
            dset = doc_art_set(r["art"])  # document level
            kidset = kontrol_id_art_set(r["kontrol_id"])  # kontrol_id level
            if not art_sets_match(req, dset, kidset):
                continue
            rows.append({**dict(r), "cos": float(s), "doc_art_set": dset, "kid_art_set": kidset})

    if not rows:
        return []

    if HAS_BM25:
        corpus_tokens = [tokens(r["text"]) for r in rows]
        bm25 = BM25Okapi(corpus_tokens)
        bm25_scores = bm25.get_scores(q_toks)
    else:
        bm25_scores = [0.0]*len(rows)

    out: List[Dict[str,Any]] = []
    for r, bm in zip(rows, bm25_scores):
        ttoks = tokens(r["text"])  # för coverage
        cov = coverage_score(q_toks, ttoks)
        lp = length_penalty(q_len, len(r["text"]))
        bm_norm = 1.0 - math.exp(-float(bm)/3.0)
        bonus = 0.0
        if ref_list:
            t = r["text"].lower()
            if any(ref.lower() in t for ref in ref_list):
                bonus += weights.bonus_ref
        # art-bonus: endast om det finns minst en gemensam art mellan "required" och (doc_set ∪ kid_set)
        req = arter_required or set()
        if req and len(req & (r["doc_art_set"] | r["kid_art_set"])) > 0:
            bonus += weights.bonus_art
        meta = weights.w_cos*float(r["cos"]) + weights.w_bm25*bm_norm + weights.w_cov*cov - weights.w_len*(1.0-lp) + bonus
        out.append({**r, "bm25": bm_norm, "cov": cov, "lenpen": lp, "meta": float(meta)})

    out.sort(key=lambda x: x["meta"], reverse=True)
    return out[:topk_return]

# =========================
# Per-nivå sammanställning
# =========================
LEVEL_ORDER = ["EUF","EUD","SFS","SVF","FÖR","ARÅD","VGL","VERIFIKAT","ANNAT"]


def pick_per_level(cands: List[Dict[str,Any]], *, threshold: float, top_docs_per_level: int = 1, top_chunks_per_doc: int = 3) -> Dict[str, Any]:
    grouped: Dict[str, Dict[int, List[Dict[str,Any]]]] = {}
    for r in cands:
        lvl = r["juridisk_niva"] or "ANNAT"
        if r["meta"] < threshold:
            continue
        grouped.setdefault(lvl, {}).setdefault(int(r["doc_id"]), []).append(r)

    result: Dict[str, Any] = {}
    for lvl in LEVEL_ORDER:
        docs = grouped.get(lvl, {})
        if not docs:
            continue
        doc_scores = []
        for did, items in docs.items():
            best = max(items, key=lambda x: x["meta"])
            doc_scores.append((did, best["meta"], best))
        doc_scores.sort(key=lambda x: x[1], reverse=True)
        chosen = []
        for did, sc, best in doc_scores[:max(1, top_docs_per_level)]:
            its = sorted(docs[did], key=lambda x: x["meta"], reverse=True)[:top_chunks_per_doc]
            chosen.append({
                "doc_id": did, "title": best["title"], "level": lvl,
                "doc_meta": float(sc), "url": best.get("url"),
                "chunks": [
                    {
                        "kontrol_id": it.get("kontrol_id"),
                        "section_ref": it.get("section_ref"),
                        "anchor_url": it.get("anchor_url") or it.get("url"),
                        "meta": float(it.get("meta",0)),
                        "text": it.get("text","")[:1200]
                    } for it in its
                ]
            })
        result[lvl] = chosen
    return result

# =========================
# Excel export helpers
# =========================
EXPORT_HEADER = [
    # EU-paket
    "Juridiskt dokument (EU)", "Sammanfattat krav (EU)", "Textutdrag (EU)", "Artikel/Paragraf + källa (EU)", "Länk (EU)", "Nivå (EU)",
    # SFS/SVF-paket
    "Juridiskt dokument (SFS/SVF)", "Sammanfattat krav (SFS/SVF)", "Textutdrag (SFS/SVF)", "Artikel/Paragraf + källa (SFS/SVF)", "Länk (SFS/SVF)", "Nivå (SFS/SVF)",
    # FÖR-paket
    "Juridiskt dokument (FÖR)", "Sammanfattat krav (FÖR)", "Textutdrag (FÖR)", "Artikel/Paragraf + källa (FÖR)", "Länk (FÖR)", "Nivå (FÖR)",
    # VGL/Verifikat-paket
    "Juridiskt dokument (VGL)", "Sammanfattat krav (VGL)", "Textutdrag (VGL)", "Artikel/Paragraf + källa (VGL)", "Länk (VGL)", "Nivå (VGL)",
    # Kandidater (3 dokument * 3 chunks)
    "Kandidat 1 – #1 chunk [score]", "Kandidat 1 – #2 chunk [score]", "Kandidat 1 – #3 chunk [score]", "Kandidat 1 – titel", "Kandidat 1 – similarity hela texten", "Kandidat 1 – länk",
    "Kandidat 2 – #1 chunk [score]", "Kandidat 2 – #2 chunk [score]", "Kandidat 2 – #3 chunk [score]", "Kandidat 2 – titel", "Kandidat 2 – similarity hela texten", "Kandidat 2 – länk",
    "Kandidat 3 – #1 chunk [score]", "Kandidat 3 – #2 chunk [score]", "Kandidat 3 – #3 chunk [score]", "Kandidat 3 – titel", "Kandidat 3 – similarity hela texten", "Kandidat 3 – länk",
]


def summarize_chunk_text(t: str) -> str:
    sents = sent_split(t)
    if sents:
        out = sents[0]
    else:
        out = t[:240]
    return out[:240]


def fill_level_cells(level_key: str, chosen: Dict[str, Any]) -> List[str]:
    if not chosen or level_key not in chosen:
        return ["", "", "", "", "", level_key]
    doc = chosen[level_key][0]
    title = doc["title"] or ""
    lvl = doc["level"] or level_key
    chunks = doc["chunks"]
    if chunks:
        ch = chunks[0]
        summa = summarize_chunk_text(ch["text"])
        section = ch.get("section_ref") or ""
        link = ch.get("anchor_url") or ""
        excerpt = ch.get("text", "")[:420]
    else:
        summa = section = link = excerpt = ""
    return [title, summa, excerpt, section, link, lvl]


def candidate_cells(all_cands: List[Dict[str,Any]], how_many_docs: int = 3, chunks_per_doc: int = 3) -> List[str]:
    per_doc: Dict[int, Dict[str,Any]] = {}
    for r in all_cands:
        did = int(r["doc_id"])
        if did not in per_doc or r["meta"] > per_doc[did]["meta"]:
            per_doc[did] = {
                "doc_id": did,
                "title": r["title"],
                "url": r.get("url") or r.get("anchor_url"),
                "meta": float(r["meta"]),
            }
    top_docs = sorted(per_doc.values(), key=lambda x: x["meta"], reverse=True)[:how_many_docs]

    cells: List[str] = []
    for d in top_docs:
        chunks = [r for r in all_cands if int(r["doc_id"])==d["doc_id"]]
        chunks.sort(key=lambda x: x["meta"], reverse=True)
        entries = []
        for ch in chunks[:chunks_per_doc]:
            entries.append(f"{(ch.get('kontrol_id') or '')} — {summarize_chunk_text(ch['text'])[:140]} [{ch['meta']:.3f}]")
        while len(entries) < chunks_per_doc:
            entries.append("")
        cells.extend(entries + [d.get("title") or "", f"{d['meta']:.3f}", d.get("url") or ""])
    while len(cells) < 18:
        cells.append("")
    return cells


def append_columns_to_excel(xlsx_bytes: bytes, rows_payload: List[List[str]], header: List[str], sheet_name: Optional[str] = None, out_path: Optional[str] = None) -> str:
    if not HAS_OPENPYXL:
        raise RuntimeError("Installera openpyxl för Excel-export")
    bio = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(bio)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    max_col = ws.max_column
    start_col = max_col + 1

    for i, h in enumerate(header):
        c = ws.cell(row=1, column=start_col + i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True)

    for r_idx, payload in enumerate(rows_payload, start=2):
        for i, val in enumerate(payload):
            c = ws.cell(row=r_idx, column=start_col + i, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    for i in range(len(header)):
        col_letter = get_column_letter(start_col + i)
        ws.column_dimensions[col_letter].width = min(64, max(18, len(header[i]) * 0.9))

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out_path = out_path or os.path.join(APP_DIR, f"export_with_candidates_{ts}.xlsx")
    wb.save(out_path)
    return out_path

# =========================
# Streamlit UI pages
# =========================

def ui_sidebar() -> str:
    st.sidebar.title("REGELKEDJA")
    with st.sidebar.expander("Status", expanded=True):
        try:
            with contextlib.closing(get_conn()) as conn:
                n_docs = conn.execute("SELECT COUNT(*) FROM documents WHERE active=1").fetchone()[0]
                n_chunks = conn.execute("SELECT COUNT(*) FROM chunks WHERE active=1").fetchone()[0]
        except Exception:
            n_docs = n_chunks = 0
        st.write(f"📄 Dokument: **{n_docs}**  🧩 Chunks: **{n_chunks}**")
        st.caption("FAISS-index " + ("✅" if os.path.isfile(FAISS_PATH) else "saknas (bygg)"))
    st.sidebar.markdown("---")
    return st.sidebar.radio("Sidor", [
        "➕ Lägg till innehåll", "🌐 Hämta via URL", "📎 Ladda upp fil",
        "🗂️ Dokument & redigering", "🏷️ Buckets & tags", "🧰 Index & sök",
        "🔗 Match & Export"
    ])


def _chunk_controls(prefix: str = "") -> ChunkOptions:
    st.markdown("#### Chunk-inställningar")
    c1,c2,c3 = st.columns(3)
    with c1:
        maxc = st.number_input(f"Max tecken / chunk {prefix}", 800, 10000, 3000, 100)
        minc = st.number_input(f"Min tecken (merge) {prefix}", 200, 4000, 1000, 50)
    with c2:
        over = st.number_input(f"Överhörning (tecken) {prefix}", 0, 2000, 220, 10)
        subs = st.number_input(f"Dela vid underrubrik över (tecken) {prefix}", 400, 10000, 2600, 100)
    with c3:
        join = st.checkbox(f"Slå ihop korta chunkar {prefix}", True)
    return ChunkOptions(int(maxc), int(over), int(subs), int(minc), bool(join))


def page_add_manual():
    st.header("➕ Lägg till manuell text")
    with st.form("add_manual"):
        c1,c2 = st.columns(2)
        with c1:
            title = st.text_input("Titel/namn")
            url   = st.text_input("Hyperlänk (valfri)")
            org   = st.text_input("Organisation", value="Jordbruksverket")
            doc_type = st.selectbox("Dokumenttyp", ["bilaga","checklista","kontrollinstruktion","avdragsvägledning","SJVFS","SVF","SFS","EUF","EUD","annat"], index=0)
        with c2:
            juridisk_niva = st.selectbox("Nivå", LEVEL_CHOICES, index=7)
            art   = st.text_input("Art (t.ex. Nötkreatur, Får, Getter, Gris – flera med komma)")
            buckets = st.text_input("Buckets (komma-separerat)", value="verifikat")
            tags    = st.text_input("Tags (komma-separerat)", value="")
        text = st.text_area("Innehåll", height=260)
        prof = st.selectbox("Chunk-profil", ["auto","vagledning","checklista","fallback"], index=0)
        opt = _chunk_controls("(manuell)")
        ok = st.form_submit_button("Spara & indexera")
        if ok:
            try:
                doc_id = insert_document(
                    title=title, url=url or None, org=org or None, doc_type=doc_type or None,
                    juridisk_niva=juridisk_niva or None, art=art or None,
                    file_bytes=None, text_input=text,
                    buckets=[x.strip() for x in buckets.split(',') if x.strip()],
                    tags=[x.strip() for x in tags.split(',') if x.strip()],
                    chunk_profile=prof, chunk_options=opt,
                )
                st.success(f"Sparat dokument #{doc_id}")
            except Exception as e:
                st.error(f"Misslyckades: {e}")


def page_fetch_url():
    st.header("🌐 Hämta via URL")
    with st.form("fetch"):
        url = st.text_input("URL")
        title = st.text_input("Titel (valfri, annars auto)")
        org = st.text_input("Organisation", value="Jordbruksverket")
        c1,c2 = st.columns(2)
        with c1:
            doc_type = st.selectbox("Dokumenttyp", ["bilaga","checklista","kontrollinstruktion","avdragsvägledning","SJVFS","SVF","SFS","EUF","EUD","annat"], index=0)
            juridisk_niva = st.selectbox("Nivå", LEVEL_CHOICES, index=7)
        with c2:
            art = st.text_input("Art (flera med komma)")
            buckets = st.text_input("Buckets", value="verifikat")
            tags = st.text_input("Tags", value="")
        prof = st.selectbox("Chunk-profil", ["auto","vagledning","checklista","fallback"], index=0)
        opt = _chunk_controls("(URL)")
        go = st.form_submit_button("Hämta & indexera")
        if go:
            try:
                txt = fetch_url(url)
                auto_title = title or (url.split('/')[-1] if url else "")
                doc_id = insert_document(
                    title=auto_title, url=url or None, org=org or None, doc_type=doc_type or None,
                    juridisk_niva=juridisk_niva or None, art=art or None,
                    file_bytes=None, text_input=txt,
                    buckets=[x.strip() for x in buckets.split(',') if x.strip()],
                    tags=[x.strip() for x in tags.split(',') if x.strip()],
                    chunk_profile=prof, chunk_options=opt,
                )
                st.success(f"Sparat dokument #{doc_id}")
            except Exception as e:
                st.error(f"Misslyckades: {e}")


def page_upload_file():
    st.header("📎 Ladda upp fil (PDF/DOCX/TXT)")
    up = st.file_uploader("Välj fil", type=["pdf","docx","txt"])
    title = st.text_input("Titel (om tomt används filnamnet/metadata)")
    url   = st.text_input("Käll-länk (valfri, behövs för klickbara #page)")
    org   = st.text_input("Organisation", value="Jordbruksverket")
    c1,c2 = st.columns(2)
    with c1:
        doc_type = st.selectbox("Dokumenttyp", ["bilaga","checklista","kontrollinstruktion","avdragsvägledning","SJVFS","SVF","SFS","EUF","EUD","annat"], index=0)
        juridisk_niva = st.selectbox("Nivå", LEVEL_CHOICES, index=7)
    with c2:
        art = st.text_input("Art (flera med komma)")
        buckets = st.text_input("Buckets", value="verifikat")
        tags = st.text_input("Tags", value="")
    prof = st.selectbox("Chunk-profil", ["auto","vagledning","checklista","fallback"], index=0)
    opt = _chunk_controls("(fil)")
    if st.button("Extrahera & indexera", disabled=(up is None)):
        if not up:
            return
        file_bytes = up.read()
        try:
            name = title or up.name
            doc_id = insert_document(
                title=name, url=url or None, org=org or None, doc_type=doc_type or None,
                juridisk_niva=juridisk_niva or None, art=art or None,
                file_bytes=file_bytes, text_input=None,
                buckets=[x.strip() for x in buckets.split(',') if x.strip()],
                tags=[x.strip() for x in tags.split(',') if x.strip()],
                chunk_profile=prof, chunk_options=opt,
            )
            st.success(f"Sparat och indexerat #{doc_id}")
        except Exception as e:
            st.error(f"Misslyckades: {e}")


def page_documents():
    st.header("🗂️ Dokument & redigering")
    with contextlib.closing(get_conn()) as conn:
        df = pd.read_sql_query(
            """
            SELECT d.doc_id, d.title, d.org, d.doc_type, d.juridisk_niva, d.art, d.url, d.created_at, d.updated_at, d.active,
                   GROUP_CONCAT(DISTINCT b.name) AS buckets,
                   GROUP_CONCAT(DISTINCT t.name) AS tags,
                   COUNT(c.chunk_id) AS chunks
            FROM documents d
            LEFT JOIN doc_buckets db ON db.doc_id=d.doc_id
            LEFT JOIN buckets b ON b.bucket_id=db.bucket_id
            LEFT JOIN doc_tags dt ON dt.doc_id=d.doc_id
            LEFT JOIN tags t ON t.tag_id=dt.tag_id
            LEFT JOIN chunks c ON c.doc_id=d.doc_id AND c.active=1
            GROUP BY d.doc_id
            ORDER BY d.doc_id DESC
            """, conn)
    st.dataframe(df, use_container_width=True, height=360)

    sel = st.selectbox("Välj dokument", ["(inget)"] + [f"#{r.doc_id} – {r.title}" for r in df.itertuples()])
    if sel != "(inget)":
        picked = int(sel.split()[0].replace("#",""))
        with contextlib.closing(get_conn()) as conn:
            row = conn.execute("SELECT * FROM documents WHERE doc_id=?", (picked,)).fetchone()
            cur_b = [r[0] for r in conn.execute("SELECT b.name FROM buckets b JOIN doc_buckets db ON db.bucket_id=b.bucket_id WHERE db.doc_id=?", (picked,)).fetchall()]
            cur_t = [r[0] for r in conn.execute("SELECT t.name FROM tags t JOIN doc_tags dt ON dt.tag_id=t.tag_id WHERE dt.doc_id=?", (picked,)).fetchall()]
        st.subheader(row["title"]) 
        c1,c2 = st.columns(2)
        with c1:
            new_title = st.text_input("Titel", value=row["title"] or "")
            new_url   = st.text_input("URL", value=row["url"] or "")
            new_org   = st.text_input("Organisation", value=row["org"] or "")
            new_type  = st.text_input("Typ", value=row["doc_type"] or "")
        with c2:
            new_niva  = st.text_input("Nivå", value=row["juridisk_niva"] or "")
            new_art   = st.text_input("Art", value=row["art"] or "")
            new_buckets = st.text_input("Buckets", value=", ".join(cur_b))
            new_tags    = st.text_input("Tags", value=", ".join(cur_t))
        d1,d2,d3,d4 = st.columns(4)
        if d1.button("💾 Spara"):
            with contextlib.closing(get_conn()) as conn:
                conn.execute("UPDATE documents SET title=?, url=?, org=?, doc_type=?, juridisk_niva=?, art=?, updated_at=? WHERE doc_id=?",
                             (new_title.strip(), (new_url.strip() or None), new_org.strip() or None, new_type.strip() or None, new_niva.strip() or None, new_art.strip() or None, now_iso(), picked))
                conn.execute("DELETE FROM doc_buckets WHERE doc_id=?", (picked,))
                for b in [x.strip() for x in new_buckets.split(',') if x.strip()]:
                    bid = upsert_bucket(conn, b)
                    if bid:
                        conn.execute("INSERT OR IGNORE INTO doc_buckets(doc_id, bucket_id) VALUES (?,?)", (picked, bid))
                conn.execute("DELETE FROM doc_tags WHERE doc_id=?", (picked,))
                for t in [x.strip() for x in new_tags.split(',') if x.strip()]:
                    tid = upsert_tag(conn, t)
                    if tid:
                        conn.execute("INSERT OR IGNORE INTO doc_tags(doc_id, tag_id) VALUES (?,?)", (picked, tid))
                conn.commit()
            st.success("Uppdaterat")
        if d2.button("🗑️ Inaktivera"):
            with contextlib.closing(get_conn()) as conn:
                conn.execute("UPDATE documents SET active=0, updated_at=? WHERE doc_id=?", (now_iso(), picked))
                conn.execute("UPDATE chunks SET active=0, updated_at=? WHERE doc_id=?", (now_iso(), picked))
                conn.commit()
            st.warning("Dokument inaktiverat. Kör \"Bygg om index\" under Index & sök.")
        if d3.button("👀 Förhandsvisa 3 chunks"):
            with contextlib.closing(get_conn()) as conn:
                dfc = pd.read_sql_query("SELECT ordinal, kontrol_id, section_ref, substr(text,1,500) AS preview, anchor_url FROM chunks WHERE doc_id=? AND active=1 ORDER BY ordinal LIMIT 3", conn, params=(picked,))
            st.dataframe(dfc, use_container_width=True, height=200)
        if d4.button("🔄 Re-chunka (inkluderande)"):
            opt = _chunk_controls("(re-chunk)")
            profile = st.selectbox("Profil", ["auto","vagledning","checklista","fallback"], index=0, key=f"reprof_{picked}")
            if st.button("Kör re-chunk nu", key=f"rebtn_{picked}"):
                try:
                    n = rechunk_document(picked, profile=profile, opt=opt)
                    st.success(f"Ny chunkning klar: {n} chunks")
                except Exception as e:
                    st.error(f"Fel: {e}")


def page_buckets_tags():
    st.header("🏷️ Buckets & tags")
    with contextlib.closing(get_conn()) as conn:
        df_b = pd.read_sql_query("SELECT bucket_id, name FROM buckets ORDER BY name", conn)
        df_t = pd.read_sql_query("SELECT tag_id, name FROM tags ORDER BY name", conn)
    c1,c2 = st.columns(2)
    with c1:
        st.subheader("Buckets")
        st.dataframe(df_b, height=240, use_container_width=True)
        new_b = st.text_input("Ny bucket")
        if st.button("Skapa bucket") and new_b.strip():
            with contextlib.closing(get_conn()) as conn:
                upsert_bucket(conn, new_b.strip()); conn.commit()
            st.experimental_rerun()
    with c2:
        st.subheader("Tags")
        st.dataframe(df_t, height=240, use_container_width=True)
        new_t = st.text_input("Ny tag")
        if st.button("Skapa tag") and new_t.strip():
            with contextlib.closing(get_conn()) as conn:
                upsert_tag(conn, new_t.strip()); conn.commit()
            st.experimental_rerun()


def page_index_tools():
    st.header("🧰 Index & sök")
    model = get_model()
    if st.button("🔁 Bygg om FAISS från DB"):
        try:
            n = rebuild_faiss_from_db(model)
            st.success(f"Index klart: {n} chunks")
        except Exception as e:
            st.error(f"Fel: {e}")
    st.markdown("---")
    q = st.text_input("Snabbsök (fråga)")
    arter_sel = set(parse_art_list(",".join(st.multiselect("Filter: välj art(er)", ARTER_CHOICES, default=[]))))
    topk = st.slider("Top-K return", 10, 200, 50)
    if q:
        rows = hybrid_search(q, arter_required=arter_sel or None, topk_return=int(topk))
        if not rows:
            st.info("Inga kandidater")
        else:
            st.dataframe(pd.DataFrame([
                {
                    "nivå": r["juridisk_niva"], "titel": r["title"], "kontroll": r["kontrol_id"],
                    "section": r["section_ref"], "meta": round(r["meta"],3), "cos": round(r["cos"],3),
                    "bm25": round(r["bm25"],3), "cov": round(r["cov"],3), "länk": r.get("anchor_url")
                } for r in rows
            ]), use_container_width=True, height=420)


def page_match_export():
    st.header("🔗 Match & Export")
    up = st.file_uploader("Excel (.xlsx)", type=["xlsx"])
    sheet = st.text_input("Bladnamn (tomt = aktivt första)", "")
    text_cols = st.text_input("Textkolumn(er) för fråga (komma-separerat)", value="Tema")
    art_col = st.text_input("Art-kolumn (kan innehålla flera, separerade med komma)", value="")
    art_defaults = set(parse_art_list(",".join(st.multiselect("Default art(er) om cell saknas)", ARTER_CHOICES, default=[]))))

    st.markdown("**Ranknings-inställningar**")
    w_cos = st.slider("w_cos (embedding)", 0.0, 1.0, 0.55, 0.05)
    w_bm25= st.slider("w_bm25 (BM25)", 0.0, 1.0, 0.25, 0.05)
    w_cov = st.slider("w_cov (coverage)", 0.0, 1.0, 0.10, 0.05)
    w_len = st.slider("w_len penalty", 0.0, 1.0, 0.10, 0.05)
    b_ref = st.slider("bonus ref-match", 0.0, 0.5, 0.12, 0.01)
    b_art = st.slider("bonus art-match", 0.0, 0.5, 0.05, 0.01)

    top_docs = st.slider("Top dokument per nivå", 1, 3, 1)
    top_chunks = st.slider("Top chunks per dokument", 1, 5, 3)
    thresh = st.slider("Meta-score threshold", 0.0, 1.0, 0.60, 0.01)
    k_dense = st.slider("FAISS TopK (dense)", 50, 1000, 400, 50)

    def art_set_for_row(row: pd.Series) -> Set[str]:
        if art_col and art_col in row.index and str(row.get(art_col, "")).strip():
            return parse_art_list(str(row.get(art_col)))
        return set(art_defaults)

    if up is not None and st.button("🚀 Matcha & förhandsvisa (10 första)"):
        try:
            df_in = pd.read_excel(up, sheet_name=sheet if sheet else None)
            if isinstance(df_in, dict):
                df_in = list(df_in.values())[0]
            cols = [c.strip() for c in text_cols.split(',') if c.strip()]
            for c in cols:
                if c not in df_in.columns:
                    st.error(f"Hittar inte kolumn: {c}")
                    return
            W = Weights(w_cos, w_bm25, w_cov, w_len, b_ref, b_art)
            previews = []
            for i in range(min(10, len(df_in))):
                row = df_in.iloc[i]
                parts = [str(row[c]) for c in cols if str(row.get(c,""))]
                query = " | ".join(parts)
                arter_req = art_set_for_row(row)
                cands = hybrid_search(query, arter_required=arter_req or None, topk_dense=int(k_dense), weights=W)
                per_level = pick_per_level(cands, threshold=float(thresh), top_docs_per_level=int(top_docs), top_chunks_per_doc=int(top_chunks))
                row_cells = []
                for key in ["EUF","SFS","FÖR","VGL"]:
                    row_cells.extend(fill_level_cells(key, per_level))
                row_cells.extend(candidate_cells(cands, how_many_docs=3, chunks_per_doc=3))
                previews.append(row_cells)
            prev_df = pd.DataFrame(previews, columns=(
                ["EU – Juridiskt dokument","EU – Sammanfattat krav","EU – Textutdrag","EU – §","EU – Länk","EU – Nivå"]+
                ["SFS/SVF – Juridiskt dokument","SFS/SVF – Sammanfattat krav","SFS/SVF – Textutdrag","SFS/SVF – §","SFS/SVF – Länk","SFS/SVF – Nivå"]+
                ["FÖR – Juridiskt dokument","FÖR – Sammanfattat krav","FÖR – Textutdrag","FÖR – §","FÖR – Länk","FÖR – Nivå"]+
                ["VGL – Juridiskt dokument","VGL – Sammanfattat krav","VGL – Textutdrag","VGL – §","VGL – Länk","VGL – Nivå"]+
                ["K1-#1","K1-#2","K1-#3","K1-titel","K1-docscore","K1-länk",
                 "K2-#1","K2-#2","K2-#3","K2-titel","K2-docscore","K2-länk",
                 "K3-#1","K3-#2","K3-#3","K3-titel","K3-docscore","K3-länk"]
            ))
            st.dataframe(prev_df, use_container_width=True, height=420)
        except Exception as e:
            st.error(f"Fel: {e}")

    if up is not None and st.button("💾 Exportera ny Excel (hela filen)"):
        try:
            up.seek(0)
            orig_bytes = up.read()
            df_in_all = pd.read_excel(io.BytesIO(orig_bytes), sheet_name=sheet if sheet else None)
            if isinstance(df_in_all, dict):
                if sheet and sheet in df_in_all:
                    df_in = df_in_all[sheet]
                else:
                    first_name = list(df_in_all.keys())[0]
                    df_in = df_in_all[first_name]
            else:
                df_in = df_in_all
            cols = [c.strip() for c in text_cols.split(',') if c.strip()]
            for c in cols:
                if c not in df_in.columns:
                    st.error(f"Hittar inte kolumn: {c}")
                    return
            W = Weights(w_cos, w_bm25, w_cov, w_len, b_ref, b_art)
            payload_rows: List[List[str]] = []
            for i in range(len(df_in)):
                row = df_in.iloc[i]
                parts = [str(row[c]) for c in cols if str(row.get(c,""))]
                query = " | ".join(parts)
                arter_req = art_set_for_row(row)
                cands = hybrid_search(query, arter_required=arter_req or None, topk_dense=int(k_dense), weights=W)
                per_level = pick_per_level(cands, threshold=float(thresh), top_docs_per_level=int(top_docs), top_chunks_per_doc=int(top_chunks))
                row_cells = []
                for key in ["EUF","SFS","FÖR","VGL"]:
                    row_cells.extend(fill_level_cells(key, per_level))
                row_cells.extend(candidate_cells(cands, how_many_docs=3, chunks_per_doc=3))
                payload_rows.append(row_cells)
            out = append_columns_to_excel(orig_bytes, payload_rows, EXPORT_HEADER, sheet_name=sheet or None)
            st.success(f"Export klar: {out}")
            st.download_button("Ladda ner filen", data=open(out,'rb').read(), file_name=os.path.basename(out))
        except Exception as e:
            st.error(f"Exportfel: {e}")

# =========================
# Main
# =========================

def main():
    st.set_page_config(page_title="REGELKEDJA — ingest/match/export (v3)", layout="wide")
    ensure_schema()
    page = ui_sidebar()
    if page == "➕ Lägg till innehåll":
        page_add_manual()
    elif page == "🌐 Hämta via URL":
        page_fetch_url()
    elif page == "📎 Ladda upp fil":
        page_upload_file()
    elif page == "🗂️ Dokument & redigering":
        page_documents()
    elif page == "🏷️ Buckets & tags":
        page_buckets_tags()
    elif page == "🧰 Index & sök":
        page_index_tools()
    else:
        page_match_export()

if __name__ == "__main__":
    main()
